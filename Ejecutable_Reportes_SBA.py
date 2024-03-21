import tkinter as tk
import translate
from tkinter import messagebox
#UTILIZA ESTA PARTE DEL CODIGO PARA INSTALAR TODOS LOS COMPONENTES NECESARIOS:
#py -m pip install pillow
#py -m pip install openpyxl
#py -m pip install natsort
#py -m pip install tk        
#py -m pip install translate
#Esta linea inicia la rutina y todos los comandos necesarios para la ejecución del código
#pyinstaller --onefile -w Ejecutable_Reportes_SBA.py
from PIL import JpegImagePlugin
from decimal import Decimal
import openpyxl
from openpyxl.styles.borders import Border, Side
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker 
from openpyxl.drawing.xdr import XDRPoint2D, XDRPositiveSize2D
from openpyxl.utils.units import pixels_to_EMU, cm_to_EMU
JpegImagePlugin._getmp = lambda: None
import glob
import os
from natsort import natsorted


#Esta linea define los bordes de las fotos a ingresar
thick_border = Border(left=Side(style='thick'), 
                     right=Side(style='thick'), 
                     top=Side(style='thick'), 
                     bottom=Side(style='thick'))

half_thick = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thick'), 
                     bottom=Side(style='thick'))


from translate import Translator

print(openpyxl.drawing.image.PILImage)



def translate_phrase(phrase, target_language='en', correo_usuario='gidrcmyk@gmail.com'):
    translator= Translator(to_lang=target_language,from_lang='es',provider='mymemory', email=correo_usuario)
    translation = translator.translate(phrase)
    return translation

class MyGUI:
    def __init__(self):
        self.root =tk.Tk(className="-REPORTES DD SBA-")
        #ENCABEZADO
        self.label0=tk.Label(self.root,text="Reportes DD SBA",font=('Arial',12))
        self.label0.pack(padx=10,pady=10)

        #INSERTAR CORREO
        self.username_frame=tk.Frame(self.root)
        self.username_frame.pack(side='top',pady=10)
        self.label0a=tk.Label(self.username_frame,text="Correo:",font=('Arial',12))
        self.label0a.pack(side="left",padx=40)
        self.entry1=tk.Entry(self.username_frame,text="",font=('Arial',12))
        self.entry1.pack(side="left",padx=40)
        self.Btn_correo=tk.Button(self.username_frame,text="Enviar",font=('Arial',12),command=self.guardar_correo)
        self.Btn_correo.pack(side="left",padx=5)

        self.sendcorreo_frame=tk.Frame(self.root)
        self.sendcorreo_frame.pack(side='top',pady=10)
        self.lbl_correo=tk.Label(self.sendcorreo_frame,text="-USUARIO-",font=('Arial',8))
        self.lbl_correo.pack(side="left",padx=35)

        #BUSCAR DOCUMENTO BASE DE REPORTE
        self.BaseDoc_frame=tk.Frame(self.root)
        self.BaseDoc_frame.pack(side='top',pady=10)
        self.label0=tk.Label(self.BaseDoc_frame,text="DOCUMENTO BASE",font=('Arial',12))
        self.label0.pack(side="top",padx=10,pady=10)
        self.label1=tk.Label(self.BaseDoc_frame,text="Excel Base:",font=('Arial',12))
        self.label1.pack(side="left",padx=10)
        self.BaseDoc_path=tk.Entry(self.BaseDoc_frame,state='disabled',font=('Arial',12))
        self.BaseDoc_path.pack(side="left",padx=40)
        self.Btn_BaseDoc=tk.Button(self.BaseDoc_frame,text="...",font=('Arial',12),command=self.buscar_doc_Breporte)
        self.Btn_BaseDoc.pack(side="left",padx=5)
        #BUSCAR DOCUMENTO DE BASE DE DATOS
        self.DB_frame=tk.Frame(self.root)
        self.DB_frame.pack(side='top',pady=10)
        self.label2=tk.Label(self.DB_frame,text="BASE DE DATOS (opcional)",font=('Arial',12))
        self.label2.pack(side="top",padx=10,pady=10)
        self.label3=tk.Label(self.DB_frame,text="DB:",font=('Arial',12))
        self.label3.pack(side="left",padx=35)
        self.DB_path=tk.Entry(self.DB_frame,state='disabled',font=('Arial',12))
        self.DB_path.pack(side="left",padx=40)
        self.Btn_DB=tk.Button(self.DB_frame,text="...",font=('Arial',12),command=self.buscar_doc_BDatos)
        self.Btn_DB.pack(side="left",padx=5)
        self.codsite_frame=tk.Frame(self.root)
        self.codsite_frame.pack(pady=10)
        self.label4=tk.Label(self.codsite_frame,text="Código SBA:",font=('Arial',12))
        self.label4.pack(side="left",padx=20)
        self.codsite_entry=tk.Entry(self.codsite_frame,text="",font=('Arial',12))
        self.codsite_entry.pack(side="left",padx=40)
        self.Btn_codsite=tk.Button(self.codsite_frame,text="Enviar",font=('Arial',12),command=self.guardar_sitio)
        self.Btn_codsite.pack(side="left",padx=5)

        self.lblsite_frame=tk.Frame(self.root)
        self.lblsite_frame.pack(side='top',pady=10)
        self.lbl_codsite=tk.Label(self.lblsite_frame,text="-COD SITE-",font=('Arial',8))
        self.lbl_codsite.pack(side="left",padx=35)


        #BUSCAR CARPETA DE REPORTE
        #self.button3 =tk.Button(self.root, text="Buscar carpeta con las imágenes ordenadas", font=('Arial',12),command=self.buscar_carpeta)
        self.FotosReporte_frame=tk.Frame(self.root)
        self.FotosReporte_frame.pack(side='left',pady=10)
        self.label5=tk.Label(self.FotosReporte_frame,text="CARPETA CON FOTOS DEL REPORTE",font=('Arial',12))
        self.label5.pack(side="top",padx=10,pady=10)
        self.label6=tk.Label(self.FotosReporte_frame,text="Carpeta:",font=('Arial',12))
        self.label6.pack(side="left",padx=20)
        self.FotosReporte_path=tk.Entry(self.FotosReporte_frame,state='disabled',font=('Arial',12))
        self.FotosReporte_path.pack(side="left",padx=40)
        self.Btn_FotosReporte=tk.Button(self.FotosReporte_frame,text="...",font=('Arial',12),command=self.buscar_carpeta)
        self.Btn_FotosReporte.pack(side="left",padx=5)


        #EJECUTAR REPORTE
        self.EjecutarReporte_frame=tk.Frame(self.root)
        self.EjecutarReporte_frame.pack(side='left',pady=10)
        self.label7=tk.Label(self.EjecutarReporte_frame,text="EJECUCIÓN DE REPORTE",font=('Arial',12))
        self.label7.pack(side="top",padx=10,pady=10)
        self.Btn_Ejecucion=tk.Button(self.EjecutarReporte_frame,text="Ejecutar",font=('Arial',12),command=self.ejecutar)
        self.Btn_Ejecucion.pack(side="top",padx=5)
        self.lbl_Ejecucion=tk.Label(self.EjecutarReporte_frame,text="-Sin Ejecutar-",font=('Arial',8))
        self.lbl_Ejecucion.pack(side="top",padx=35)

        #BUSCAR DOCUMENTO DE SALIDA 1_ESPAÑOL
        self.Traduccion_frame=tk.Frame(self.root)
        self.Traduccion_frame.pack(side='bottom',pady=10)
        self.Btn_Traduccion=tk.Button(self.Traduccion_frame,text="Traducir",font=('Arial',12),command=self.traducir)
        self.Btn_Traduccion.pack(side="top",padx=5)
        self.lbl_Traduccion=tk.Label(self.Traduccion_frame,text="-Sin Ejecutar-",font=('Arial',8))
        self.lbl_Traduccion.pack(side="top",padx=35)

        self.PorTraducir_frame=tk.Frame(self.root)
        self.PorTraducir_frame.pack(side='bottom',pady=10)
        self.label8=tk.Label(self.PorTraducir_frame,text="DOCUMENTO PARA TRADUCCIÓN",font=('Arial',12))
        self.label8.pack(side="top",padx=10,pady=10)
        self.label9=tk.Label(self.PorTraducir_frame,text="Excel Base:",font=('Arial',12))
        self.label9.pack(side="left",padx=10)
        self.PorTraducir_path=tk.Entry(self.PorTraducir_frame,state='disabled',font=('Arial',12))
        self.PorTraducir_path.pack(side="left",padx=40)
        self.Btn_PorTraducir=tk.Button(self.PorTraducir_frame,text="...",font=('Arial',12),command=self.buscar_doc_Reportelisto)
        self.Btn_PorTraducir.pack(side="left",padx=5)



        self.root.mainloop()
    def guardar_correo(self):
        pass
        self.lbl_correo.config(text=self.entry1.get())
    def guardar_sitio(self):
        pass
        self.lbl_codsite.config(text=self.codsite_entry.get())
    def buscar_carpeta(self):
        pass
        import tkinter
        import tkinter.filedialog as tkFileDialog
        root = tkinter.Tk()
        root.lift()
        root.attributes('-topmost',True)
        root.after_idle(root.attributes,'-topmost',False)
        root.withdraw()
        folder_path = tkFileDialog.askdirectory()
        self.FotosReporte_path.configure(state='normal')
        self.FotosReporte_path.insert(0,folder_path)
        self.FotosReporte_path.configure(state='disabled')
        print(folder_path)
        root.destroy()

    def buscar_doc_Breporte(self):
        pass
        #SELECCIONAR EXCEL BASE
        import tkinter
        import tkinter.filedialog as tkFileDialog

        root = tkinter.Tk()

        root.lift()
        root.attributes('-topmost',True)
        root.after_idle(root.attributes,'-topmost',False)
        root.withdraw()
        file_path = tkFileDialog.askopenfile()
        self.BaseDoc_path.configure(state='normal')
        self.BaseDoc_path.insert(0,file_path.name)
        self.BaseDoc_path.configure(state='disabled')
        print(file_path.name)
        root.destroy()
    def buscar_doc_BDatos(self):
        pass
        #SELECCIONAR EXCEL BASE
        import tkinter
        import tkinter.filedialog as tkFileDialog

        root = tkinter.Tk()

        root.lift()
        root.attributes('-topmost',True)
        root.after_idle(root.attributes,'-topmost',False)
        root.withdraw()
        file_path = tkFileDialog.askopenfile()
        self.DB_path.configure(state='normal')
        self.DB_path.insert(0,file_path.name)
        self.DB_path.configure(state='disabled')
        print(file_path.name)
        root.destroy()
    def buscar_doc_Reportelisto(self):
        pass
        #SELECCIONAR EXCEL BASE
        import tkinter
        import tkinter.filedialog as tkFileDialog

        root = tkinter.Tk()

        root.lift()
        root.attributes('-topmost',True)
        root.after_idle(root.attributes,'-topmost',False)
        root.withdraw()
        file_path = tkFileDialog.askopenfile()
        self.PorTraducir_path.configure(state='normal')
        self.PorTraducir_path.insert(0,file_path.name)
        self.PorTraducir_path.configure(state='disabled')
        print(file_path.name)
        root.destroy()
    def ejecutar(self):
        pass
        self.lbl_Ejecucion.config(text="-En ejecución-")

        #%%%%%%%%%%%%%%%%%EJECUCIÓN PARA EL REPORTE%%%%%%%%%%%%%%%%%%%%%%%%%%
        #ABRIR EL DOCUMENTO DE EXCEL
        file_path=self.BaseDoc_path.get()
        folder_path=self.FotosReporte_path.get()
        correo_usuario=self.lbl_correo.cget("text")

        workbook = openpyxl.load_workbook(file_path)
        p2e = pixels_to_EMU
        #-------%%%%%%%HOJA 1, CELDA DE INSERCIÓN I25, ANCHO 29.89, ALTO 286.2
        ws1=workbook["1"]
        dir1=folder_path+"/1"
        Ancho1=29.89*7.25
        Alto1=286*4/3
        images=[]
        for filename in natsorted(glob.glob(dir1+"/**/*jpeg",recursive=True)):
            images.append(filename)
        
        total=len(images)

        img1=Image(images[0])
        hpercent = Alto1/img1.height
        wsize=int(img1.width/hpercent)
        coloff=0
        filoff=0
        if wsize>Ancho1:
            wpercent=(Ancho1/img1.width)
            hsize =int(img1.height*wpercent)
            img1.width=Ancho1
            img1.height=hsize
            filoff=p2e((Alto1-hsize)/2)
        else:
            img1.width=wsize
            img1.height=Alto1
            coloff=p2e((Ancho1-wsize)/2)

        img1.quality=100
        h,w=img1.height,img1.width
        size = XDRPositiveSize2D(p2e(w), p2e(h))
        ws1.add_image(img1,"I25")
        #-------%%%%%%%HOJA TOWER INF
        filainicial=7
        altoenfilas=8
        tituloenfilas=0
        separacionenfilas=0
        columnas_iniciales=[7]
        CantColumnas=len(columnas_iniciales)
        TotalSaltoFila=altoenfilas+tituloenfilas+separacionenfilas
        worksheet=workbook["TOWER INF."]
        dir1=folder_path+"/TOWER INF"
        basewidth=39*7.25
        baseheight=147*4/3
        images = []
        anclajes = []
        anc_fila=[]
        anc_col=[]
        for filename in natsorted(glob.glob(dir1+"/**/*jpeg",recursive=True)):
            images.append(filename)
        total=len(images)
        CantFilas=round(total/CantColumnas)
        FinalDeFilas=filainicial+TotalSaltoFila*CantFilas+1

        for row in range(filainicial,FinalDeFilas,TotalSaltoFila):
                for col in columnas_iniciales:
                    col_letter = get_column_letter(col)
                    anclaje = col_letter + str(row)
                    anc_fila.append(row)
                    anc_col.append(col)
                    anclajes.append(anclaje)
        for index, image in enumerate(images):
            img=Image(image)
            hpercent = (float(baseheight) / float(img.height))
            wsize = int((float(img.width) * float(hpercent)))
            coloff=0
            filoff=0
            if wsize>basewidth:
                wpercent = (float(basewidth) / float(img.width))
                hsize = int((float(img.height) * float(wpercent)))
                img.width=basewidth
                img.height=hsize
                filoff=p2e((baseheight-hsize)/2)

            else:
                img.width=wsize
                img.height=baseheight
                coloff=p2e((basewidth-wsize)/2)

            img.quality=50
            h, w = img.height, img.width
            size = XDRPositiveSize2D(p2e(w), p2e(h))
            marker = AnchorMarker(col=anc_col[index]-1, colOff=coloff, row=anc_fila[index]-1, rowOff=filoff)
            img.anchor = OneCellAnchor(_from=marker, ext=size)
            worksheet.add_image(img)
        #-------%%%%%%%HOJA COAX_MAP
        filainicial=10
        altoenfilas=1
        tituloenfilas=0
        separacionenfilas=1
        columnas_iniciales=[1,6]
        CantColumnas=len(columnas_iniciales)
        TotalSaltoFila=altoenfilas+tituloenfilas+separacionenfilas
        worksheet=workbook["COAX MAP"]
        dir1=folder_path+"/COAX MAP"
        basewidth=45*7.25
        baseheight=180*4/3
        images = []
        anclajes = []
        anc_fila=[]
        anc_col=[]
        for filename in natsorted(glob.glob(dir1+"/**/*jpeg",recursive=True)):
            images.append(filename)
        total=len(images)
        CantFilas=round(total/CantColumnas)
        FinalDeFilas=filainicial+TotalSaltoFila*CantFilas+1

        for row in range(filainicial,FinalDeFilas,TotalSaltoFila):
                for col in columnas_iniciales:
                    col_letter = get_column_letter(col)
                    anclaje = col_letter + str(row)
                    anc_fila.append(row)
                    anc_col.append(col)
                    anclajes.append(anclaje)
        for index, image in enumerate(images):
            img=Image(image)
            hpercent = (float(baseheight) / float(img.height))
            wsize = int((float(img.width) * float(hpercent)))
            coloff=0
            filoff=0
            if wsize>basewidth:
                wpercent = (float(basewidth) / float(img.width))
                hsize = int((float(img.height) * float(wpercent)))
                img.width=basewidth
                img.height=hsize
                filoff=p2e((baseheight-hsize)/2)

            else:
                img.width=wsize
                img.height=baseheight
                coloff=p2e((basewidth-wsize)/2)

            img.quality=100
            h, w = img.height, img.width
            size = XDRPositiveSize2D(p2e(w), p2e(h))
            marker = AnchorMarker(col=anc_col[index]-1, colOff=coloff, row=anc_fila[index]-1, rowOff=filoff)
            img.anchor = OneCellAnchor(_from=marker, ext=size)
            worksheet.add_image(img)
        #-------%%%%%%%HOJA PHOTOS
        #-------ACCESO
        filainicial=27
        altoenfilas=6
        tituloenfilas=0
        separacionenfilas=1
        columnas_iniciales=[1,4,7]
        CantColumnas=len(columnas_iniciales)
        TotalSaltoFila=altoenfilas+tituloenfilas+separacionenfilas
        worksheet=workbook["PHOTOS"]
        dir1=folder_path+"/PHOTOS/Acceso"
        basewidth=27*7.25
        baseheight=85.8*4/3
        images = []
        anclajes = []
        anc_fila=[]
        anc_col=[]
        for filename in natsorted(glob.glob(dir1+"/**/*jpeg",recursive=True)):
            images.append(filename)
        total=len(images)
        CantFilas=round(total/CantColumnas)
        FinalDeFilas=filainicial+TotalSaltoFila*CantFilas+1

        for row in range(filainicial,FinalDeFilas,TotalSaltoFila):
                for col in columnas_iniciales:
                    col_letter = get_column_letter(col)
                    anclaje = col_letter + str(row)
                    anc_fila.append(row)
                    anc_col.append(col)
                    anclajes.append(anclaje)
        for index, image in enumerate(images):
            img=Image(image)
            hpercent = (float(baseheight) / float(img.height))
            wsize = int((float(img.width) * float(hpercent)))
            coloff=0
            filoff=0
            if wsize>basewidth:
                wpercent = (float(basewidth) / float(img.width))
                hsize = int((float(img.height) * float(wpercent)))
                img.width=basewidth
                img.height=hsize
                filoff=p2e((baseheight-hsize)/2)

            else:
                img.width=wsize
                img.height=baseheight
                coloff=p2e((basewidth-wsize)/2)

            img.quality=100
            h, w = img.height, img.width
            size = XDRPositiveSize2D(p2e(w), p2e(h))
            marker = AnchorMarker(col=anc_col[index]-1, colOff=coloff, row=anc_fila[index]-1, rowOff=filoff)
            img.anchor = OneCellAnchor(_from=marker, ext=size)
            worksheet.add_image(img)   
        #-------LOCATION
        filainicial=60
        altoenfilas=6
        tituloenfilas=0
        separacionenfilas=1
        columnas_iniciales=[1,4,7]
        CantColumnas=len(columnas_iniciales)
        TotalSaltoFila=altoenfilas+tituloenfilas+separacionenfilas
        worksheet=workbook["PHOTOS"]
        dir1=folder_path+"/PHOTOS/Location"
        basewidth=27*7.25
        baseheight=85.8*4/3
        images = []
        anclajes = []
        anc_fila=[]
        anc_col=[]
        for filename in natsorted(glob.glob(dir1+"/**/*jpeg",recursive=True)):
            images.append(filename)
        total=len(images)
        CantFilas=round(total/CantColumnas)
        FinalDeFilas=filainicial+TotalSaltoFila*CantFilas+1

        for row in range(filainicial,FinalDeFilas,TotalSaltoFila):
                for col in columnas_iniciales:
                    col_letter = get_column_letter(col)
                    anclaje = col_letter + str(row)
                    anc_fila.append(row)
                    anc_col.append(col)
                    anclajes.append(anclaje)
        for index, image in enumerate(images):
            img=Image(image)
            hpercent = (float(baseheight) / float(img.height))
            wsize = int((float(img.width) * float(hpercent)))
            coloff=0
            filoff=0
            if wsize>basewidth:
                wpercent = (float(basewidth) / float(img.width))
                hsize = int((float(img.height) * float(wpercent)))
                img.width=basewidth
                img.height=hsize
                filoff=p2e((baseheight-hsize)/2)

            else:
                img.width=wsize
                img.height=baseheight
                coloff=p2e((basewidth-wsize)/2)

            img.quality=100
            h, w = img.height, img.width
            size = XDRPositiveSize2D(p2e(w), p2e(h))
            marker = AnchorMarker(col=anc_col[index]-1, colOff=coloff, row=anc_fila[index]-1, rowOff=filoff)
            img.anchor = OneCellAnchor(_from=marker, ext=size)
            worksheet.add_image(img)     
        #-------PANORAMICAS
        filainicial=135
        altoenfilas=17
        tituloenfilas=0
        separacionenfilas=1
        columnas_iniciales=[1,6]
        CantColumnas=len(columnas_iniciales)
        TotalSaltoFila=altoenfilas+tituloenfilas+separacionenfilas
        worksheet=workbook["PHOTOS"]
        dir1=folder_path+"/PHOTOS/Panoramicas"
        basewidth=45*7.25
        baseheight=224.4*4/3
        images = []
        anclajes = []
        anc_fila=[]
        anc_col=[]
        for filename in natsorted(glob.glob(dir1+"/**/*jpeg",recursive=True)):
            images.append(filename)
        total=len(images)
        CantFilas=round(total/CantColumnas)
        FinalDeFilas=filainicial+TotalSaltoFila*CantFilas+1

        for row in range(filainicial,FinalDeFilas,TotalSaltoFila):
                for col in columnas_iniciales:
                    col_letter = get_column_letter(col)
                    anclaje = col_letter + str(row)
                    anc_fila.append(row)
                    anc_col.append(col)
                    anclajes.append(anclaje)
        for index, image in enumerate(images):
            if index==6:
                break
            img=Image(image)
            hpercent = (float(baseheight) / float(img.height))
            wsize = int((float(img.width) * float(hpercent)))
            coloff=0
            filoff=0
            if wsize>basewidth:
                wpercent = (float(basewidth) / float(img.width))
                hsize = int((float(img.height) * float(wpercent)))
                img.width=basewidth
                img.height=hsize
                filoff=p2e((baseheight-hsize)/2)

            else:
                img.width=wsize
                img.height=baseheight
                coloff=p2e((basewidth-wsize)/2)

            img.quality=100
            h, w = img.height, img.width
            size = XDRPositiveSize2D(p2e(w), p2e(h))
            marker = AnchorMarker(col=anc_col[index]-1, colOff=coloff, row=anc_fila[index]-1, rowOff=filoff)
            img.anchor = OneCellAnchor(_from=marker, ext=size)
            worksheet.add_image(img) 
        #-------PANORAMICAS 2da parte
        filainicial=139
        altoenfilas=17
        tituloenfilas=0
        separacionenfilas=1
        columnas_iniciales=[1,6]
        CantColumnas=len(columnas_iniciales)
        TotalSaltoFila=altoenfilas+tituloenfilas+separacionenfilas
        worksheet=workbook["PHOTOS"]
        dir1=folder_path+"/PHOTOS/Panoramicas"
        basewidth=45*7.25
        baseheight=224.4*4/3
        images = []
        anclajes = []
        anc_fila=[]
        anc_col=[]
        for filename in natsorted(glob.glob(dir1+"/**/*jpeg",recursive=True)):
            images.append(filename)
        total=len(images)
        CantFilas=round(total/CantColumnas)
        FinalDeFilas=filainicial+TotalSaltoFila*CantFilas+1

        for row in range(filainicial,FinalDeFilas,TotalSaltoFila):
                for col in columnas_iniciales:
                    col_letter = get_column_letter(col)
                    anclaje = col_letter + str(row)
                    anc_fila.append(row)
                    anc_col.append(col)
                    anclajes.append(anclaje)
        for index, image in enumerate(images):
            if index>=6:
                img=Image(image)
                hpercent = (float(baseheight) / float(img.height))
                wsize = int((float(img.width) * float(hpercent)))
                coloff=0
                filoff=0
                if wsize>basewidth:
                    wpercent = (float(basewidth) / float(img.width))
                    hsize = int((float(img.height) * float(wpercent)))
                    img.width=basewidth
                    img.height=hsize
                    filoff=p2e((baseheight-hsize)/2)

                else:
                    img.width=wsize
                    img.height=baseheight
                    coloff=p2e((basewidth-wsize)/2)

                img.quality=100
                h, w = img.height, img.width
                size = XDRPositiveSize2D(p2e(w), p2e(h))
                marker = AnchorMarker(col=anc_col[index]-1, colOff=coloff, row=anc_fila[index]-1, rowOff=filoff)
                img.anchor = OneCellAnchor(_from=marker, ext=size)
                worksheet.add_image(img) 
        #-------SITE PICTURES 2
        filainicial=82
        altoenfilas=10
        tituloenfilas=0
        separacionenfilas=1
        columnas_iniciales=[1,4,7]
        CantColumnas=len(columnas_iniciales)
        TotalSaltoFila=altoenfilas+tituloenfilas+separacionenfilas
        worksheet=workbook["PHOTOS"]
        dir1=folder_path+"/PHOTOS/SitePictures2"
        basewidth=27*7.25
        baseheight=130*4/3
        images = []
        anclajes = []
        anc_fila=[]
        anc_col=[]
        for filename in natsorted(glob.glob(dir1+"/**/*jpeg",recursive=True)):
            images.append(filename)
        total=len(images)
        CantFilas=round(total/CantColumnas)
        FinalDeFilas=filainicial+TotalSaltoFila*CantFilas+1

        for row in range(filainicial,FinalDeFilas,TotalSaltoFila):
                for col in columnas_iniciales:
                    col_letter = get_column_letter(col)
                    anclaje = col_letter + str(row)
                    anc_fila.append(row)
                    anc_col.append(col)
                    anclajes.append(anclaje)
        for index, image in enumerate(images):

            img=Image(image)
            hpercent = (float(baseheight) / float(img.height))
            wsize = int((float(img.width) * float(hpercent)))
            coloff=0
            filoff=0
            if wsize>basewidth:
                wpercent = (float(basewidth) / float(img.width))
                hsize = int((float(img.height) * float(wpercent)))
                img.width=basewidth
                img.height=hsize
                filoff=p2e((baseheight-hsize)/2)

            else:
                img.width=wsize
                img.height=baseheight
                coloff=p2e((basewidth-wsize)/2)

            img.quality=100
            h, w = img.height, img.width
            size = XDRPositiveSize2D(p2e(w), p2e(h))
            marker = AnchorMarker(col=anc_col[index]-1, colOff=coloff, row=anc_fila[index]-1, rowOff=filoff)
            img.anchor = OneCellAnchor(_from=marker, ext=size)
            worksheet.add_image(img)  
        #-------TIERRAS
        filainicial=7
        altoenfilas=6
        tituloenfilas=0
        separacionenfilas=1
        columnas_iniciales=[1,4,7]
        CantColumnas=len(columnas_iniciales)
        TotalSaltoFila=altoenfilas+tituloenfilas+separacionenfilas
        worksheet=workbook["PHOTOS"]
        dir1=folder_path+"/PHOTOS/Tierra"
        basewidth=27*7.25
        baseheight=85.8*4/3
        images = []
        anclajes = []
        anc_fila=[]
        anc_col=[]
        for filename in natsorted(glob.glob(dir1+"/**/*jpeg",recursive=True)):
            images.append(filename)
        total=len(images)
        CantFilas=round(total/CantColumnas)
        FinalDeFilas=filainicial+TotalSaltoFila*CantFilas+1

        for row in range(filainicial,FinalDeFilas,TotalSaltoFila):
                for col in columnas_iniciales:
                    col_letter = get_column_letter(col)
                    anclaje = col_letter + str(row)
                    anc_fila.append(row)
                    anc_col.append(col)
                    anclajes.append(anclaje)
        for index, image in enumerate(images):
            img=Image(image)
            hpercent = (float(baseheight) / float(img.height))
            wsize = int((float(img.width) * float(hpercent)))
            coloff=0
            filoff=0
            if wsize>basewidth:
                wpercent = (float(basewidth) / float(img.width))
                hsize = int((float(img.height) * float(wpercent)))
                img.width=basewidth
                img.height=hsize
                filoff=p2e((baseheight-hsize)/2)

            else:
                img.width=wsize
                img.height=baseheight
                coloff=p2e((basewidth-wsize)/2)

            img.quality=100
            h, w = img.height, img.width
            size = XDRPositiveSize2D(p2e(w), p2e(h))
            marker = AnchorMarker(col=anc_col[index]-1, colOff=coloff, row=anc_fila[index]-1, rowOff=filoff)
            img.anchor = OneCellAnchor(_from=marker, ext=size)
            worksheet.add_image(img)      
        #-------%%%%%%%HOJA PHOTOS_2
        #-------EQUIPO INSTALADO
        filainicial=7
        altoenfilas=10
        tituloenfilas=0
        separacionenfilas=0
        columnas_iniciales=[1,3,5,7]
        CantColumnas=len(columnas_iniciales)
        TotalSaltoFila=altoenfilas+tituloenfilas+separacionenfilas
        worksheet=workbook["PHOTOS_2"]
        dir1=folder_path+"/PHOTOS_2/EquipoInstalado"
        basewidth=27*7.25
        baseheight=130*4/3
        images = []
        anclajes = []
        anc_fila=[]
        anc_col=[]
        for filename in natsorted(glob.glob(dir1+"/**/*jpeg",recursive=True)):
            images.append(filename)
        total=len(images)
        CantFilas=round(total/CantColumnas)
        FinalDeFilas=filainicial+TotalSaltoFila*CantFilas+1

        for row in range(filainicial,FinalDeFilas,TotalSaltoFila):
                for col in columnas_iniciales:
                    col_letter = get_column_letter(col)
                    anclaje = col_letter + str(row)
                    anc_fila.append(row)
                    anc_col.append(col)
                    anclajes.append(anclaje)
        for index, image in enumerate(images):

            img=Image(image)
            hpercent = (float(baseheight) / float(img.height))
            wsize = int((float(img.width) * float(hpercent)))
            coloff=0
            filoff=0
            if wsize>basewidth:
                wpercent = (float(basewidth) / float(img.width))
                hsize = int((float(img.height) * float(wpercent)))
                img.width=basewidth
                img.height=hsize
                filoff=p2e((baseheight-hsize)/2)

            else:
                img.width=wsize
                img.height=baseheight
                coloff=p2e((basewidth-wsize)/2)

            img.quality=100
            h, w = img.height, img.width
            size = XDRPositiveSize2D(p2e(w), p2e(h))
            marker = AnchorMarker(col=anc_col[index]-1, colOff=coloff, row=anc_fila[index]-1, rowOff=filoff)
            img.anchor = OneCellAnchor(_from=marker, ext=size)
            worksheet.add_image(img)  
        #-------INSTALACIONES COMUNES
        filainicial=56
        altoenfilas=10
        tituloenfilas=0
        separacionenfilas=1
        columnas_iniciales=[1,4,7]
        CantColumnas=len(columnas_iniciales)
        TotalSaltoFila=altoenfilas+tituloenfilas+separacionenfilas
        worksheet=workbook["PHOTOS_2"]
        dir1=folder_path+"/PHOTOS_2/InstalacionesComunes"
        basewidth=27*7.25
        baseheight=130*4/3
        images = []
        anclajes = []
        anc_fila=[]
        anc_col=[]
        for filename in natsorted(glob.glob(dir1+"/**/*jpeg",recursive=True)):
            images.append(filename)
        total=len(images)
        CantFilas=round(total/CantColumnas)
        FinalDeFilas=filainicial+TotalSaltoFila*CantFilas+1

        for row in range(filainicial,FinalDeFilas,TotalSaltoFila):
                for col in columnas_iniciales:
                    col_letter = get_column_letter(col)
                    anclaje = col_letter + str(row)
                    anc_fila.append(row)
                    anc_col.append(col)
                    anclajes.append(anclaje)
        for index, image in enumerate(images):

            img=Image(image)
            hpercent = (float(baseheight) / float(img.height))
            wsize = int((float(img.width) * float(hpercent)))
            coloff=0
            filoff=0
            if wsize>basewidth:
                wpercent = (float(basewidth) / float(img.width))
                hsize = int((float(img.height) * float(wpercent)))
                img.width=basewidth
                img.height=hsize
                filoff=p2e((baseheight-hsize)/2)

            else:
                img.width=wsize
                img.height=baseheight
                coloff=p2e((basewidth-wsize)/2)

            img.quality=100
            h, w = img.height, img.width
            size = XDRPositiveSize2D(p2e(w), p2e(h))
            marker = AnchorMarker(col=anc_col[index]-1, colOff=coloff, row=anc_fila[index]-1, rowOff=filoff)
            img.anchor = OneCellAnchor(_from=marker, ext=size)
            worksheet.add_image(img)
        #-------%%%%%%%HOJA PHOTOS_2_3
        filainicial=7
        altoenfilas=1
        tituloenfilas=1
        separacionenfilas=0
        anchoencolumnas=5
        columnas_iniciales=[1,6]
        CantColumnas=len(columnas_iniciales)
        TotalSaltoFila=altoenfilas+tituloenfilas+separacionenfilas
        worksheet=workbook["PHOTOS_2_3"]
        dir1=folder_path+"/PHOTOS_2_3"
        basewidth=45*7.25
        baseheight=145*4/3
        images = []
        anclajes = []
        anc_fila=[]
        anc_col=[]
        for filename in natsorted(glob.glob(dir1+"/**/*jpeg",recursive=True)):
            images.append(filename)
        total=len(images)
        CantFilas=round(total/CantColumnas)
        FinalDeFilas=filainicial+TotalSaltoFila*CantFilas+1
        #Modificar celdas para que se inserten las imágenes
        for row in range(filainicial,FinalDeFilas,TotalSaltoFila):
            for col in columnas_iniciales:
                for filas in range(0,altoenfilas):
                    worksheet.row_dimensions[row+filas].height = (Decimal(15.71)+Decimal(baseheight)*(Decimal(3/4)))/Decimal(altoenfilas)
                for columnas in range(0,anchoencolumnas):
                    col_letter = get_column_letter(col+columnas)
                    #worksheet.column_dimensions[col_letter].width = (basewidth*(1/7))/anchoencolumnas
                worksheet.merge_cells(start_row=row,start_column=col,end_row=row+altoenfilas-1,end_column=col+anchoencolumnas-1)
                worksheet.merge_cells(start_row=row+altoenfilas,start_column=col,end_row=row+altoenfilas,end_column=col+anchoencolumnas-2)

        for row in range(filainicial,FinalDeFilas,TotalSaltoFila):
                for col in columnas_iniciales:
                    col_letter = get_column_letter(col)
                    anclaje = col_letter + str(row)
                    anc_fila.append(row)
                    anc_col.append(col)
                    anclajes.append(anclaje)
        for index, image in enumerate(images):
            img=Image(image)
            hpercent = (float(baseheight) / float(img.height))
            wsize = int((float(img.width) * float(hpercent)))
            coloff=0
            filoff=0
            if wsize>basewidth:
                wpercent = (float(basewidth) / float(img.width))
                hsize = int((float(img.height) * float(wpercent)))
                img.width=basewidth
                img.height=hsize
                filoff=p2e((baseheight-hsize)/2)

            else:
                img.width=wsize
                img.height=baseheight
                coloff=p2e((basewidth-wsize)/2)

            img.quality=100
            h, w = 0.95*img.height, 0.95*img.width
            size = XDRPositiveSize2D(p2e(w), p2e(h))
            marker = AnchorMarker(col=anc_col[index]-1, colOff=coloff+100000, row=anc_fila[index]-1, rowOff=filoff+90000)
            img.anchor = OneCellAnchor(_from=marker, ext=size)
            worksheet.add_image(img) 

        if tituloenfilas==1:
            titles = []
            #obtener títulos de los nombres de las carpetas
            for title in natsorted(glob.glob(dir1+"/**/*jpeg",recursive=True)):
                titulo_original=os.path.basename(os.path.dirname(title))
                titulo_traducido=translate_phrase(titulo_original,correo_usuario=correo_usuario)
                titles.append(titulo_traducido)
            

            ws1=workbook['ISSUES']
            columna=column_index_from_string('E')
            fila_ini=11
            #escribir títulos debajo de las celdas
            for index, title in enumerate(titles):
                #worksheet.cell(row=anc_fila[index], column=anc_col[index])
                #worksheet.cell(row=anc_fila[index], column=anc_col[index]+1)
                ws1.cell(column=columna,row=fila_ini+index,value=title)
                for filas in range(altoenfilas+1):
                    for columnas in range(anchoencolumnas):
                        worksheet.cell(row=anc_fila[index]+filas, column=anc_col[index]+columnas).border=thick_border

                worksheet.cell(row=anc_fila[index]+1, column=anc_col[index], value=title)

                if anchoencolumnas>1:
                    worksheet.cell(row=anc_fila[index]+1, column=anc_col[index]+anchoencolumnas-1, value=index+1)

        BaseDatos_path=self.DB_path.get()
        Codigo_site=self.lbl_codsite.cget("text")
        if BaseDatos_path != "":
            if Codigo_site != "":
                COD_CLIENTE=Codigo_site
                workbookBD = openpyxl.load_workbook(BaseDatos_path, data_only=True)
                wsbd=workbookBD["Sheet1"]

                items=[]

                for fila in wsbd['C4:C71']:
                    for cell in fila:
                        if cell.value==COD_CLIENTE:
                            for item in range(3,42):
                                items.append(wsbd.cell(row=cell.row, column=item).value)
                workbookBD.close()

                wsbd=workbook["1"]
                wsbd1_2=workbook["1_2"]
                wsbdt_i=workbook["TOWER INF."]

                wsbd['E10']=items[0]
                wsbd['E9']=items[1]

                if items[3]=='Greenfield':
                    wsbd['F28']='√'
                    wsbd['F29']='N/A'
                else:
                    wsbd['F28']='N/A'
                    wsbd['F29']='√'
                wsbd['E11']=items[6]
                wsbd['E13']=items[7]
                wsbd['E12']=items[8]
                wsbd['K5']=items[12]
                wsbd['E17']=items[13]
                wsbd['E20']=items[14]
                wsbd['F23']=items[15]
                wsbd['F24']=items[16]
                wsbd['F25']=items[17]
                wsbd['F32']=items[18]
                wsbd['F30']=items[19]
                wsbd['F31']=items[20]
                wsbd['F35']=items[21]
                wsbd['F36']=items[22]
                wsbd['F46']=items[23]
                wsbdt_i['E7']=items[23]
                wsbd['F45']=items[24]
                wsbdt_i['E13']=items[26]
                wsbd['F50']=items[27]
                wsbd['F51']=items[28]
                wsbd1_2['G8']=items[29]
                wsbd1_2['G9']=items[30]
                wsbd1_2['G10']=items[31]
                wsbd1_2['G11']=items[32]
                wsbd1_2['G12']=items[33]
                wsbd1_2['G13']=items[34]
                wsbd1_2['G14']=items[35]
                wsbd1_2['G17']=items[36]
                wsbd1_2['G18']=items[37]
                wsbd1_2['G19']=items[38]
        #GUARDARLO EN LA CARPETA DE DESTINO
        workbook.save(folder_path+'/Reporte Fotográfico.xlsx')
        #CERRAR EL DOCUMENTO DE EXCEL
        workbook.close()
        self.lbl_Ejecucion.config(text="-Listo-")
    def traducir(self):
        pass
        self.lbl_Traduccion.config(text="-En ejecución-")

                #%%%%%%%%%%%%%%%%%TRADUCCIONES%%%%%%%%%%%%%%%%%%%%%%%%%%
        directorio=self.FotosReporte_path.get()
        file_traducir=self.PorTraducir_path.get()
        user_mail=self.lbl_correo.cget("text")
        workbook = openpyxl.load_workbook(file_traducir)
        #-------HOJA 1
        ws1=workbook['1']
        columna=column_index_from_string('F')
        fila_ini=32
        fila_fin=51
        for row in ws1.iter_rows(min_row=fila_ini, max_col=columna, max_row=fila_fin,min_col=columna):
            for cell in row:
                textoInicial=cell.value
                if textoInicial!="":
                    #print(textoInicial)
                    #time.sleep(1)
                    try:
                        Texto_Traducido=translate_phrase(textoInicial,correo_usuario=user_mail)
                        if Texto_Traducido!="None":
                            cell.value=Texto_Traducido
                    except:
                        pass
        #-------HOJA 1_2
        ws1=workbook['1_2']
        columna=column_index_from_string('G')
        fila_ini=17
        fila_fin=20
        for row in ws1.iter_rows(min_row=fila_ini, max_col=columna, max_row=fila_fin,min_col=columna):
            for cell in row:
                textoInicial=cell.value
                if textoInicial!="":
                    #print(textoInicial)
                    #time.sleep(1)
                    try:
                        Texto_Traducido=translate_phrase(textoInicial,correo_usuario=user_mail)
                        if Texto_Traducido!="None":
                            cell.value=Texto_Traducido
                    except:
                        pass
        #-------HOJA ANNEX J REP.
        ws1=workbook['ANNEX J REP.']
        columna=column_index_from_string('H')
        fila_ini=15
        fila_fin=53
        for row in ws1.iter_rows(min_row=fila_ini, max_col=columna, max_row=fila_fin,min_col=columna):
            for cell in row:
                textoInicial=cell.value
                if textoInicial!="":
                    #print(textoInicial)
                    #time.sleep(1)
                    try:
                        Texto_Traducido=translate_phrase(textoInicial,correo_usuario=user_mail)
                        if Texto_Traducido!="None":
                            cell.value=Texto_Traducido
                    except:
                        pass
        #-------HOJA ANNEX J REP.2
        ws1=workbook['ANNEX J REP.2']
        columna=column_index_from_string('L')
        fila_ini=15
        fila_fin=43
        for row in ws1.iter_rows(min_row=fila_ini, max_col=columna, max_row=fila_fin,min_col=columna):
            for cell in row:
                textoInicial=cell.value
                if textoInicial!="":
                    #print(textoInicial)
                    #time.sleep(1)
                    try:
                        Texto_Traducido=translate_phrase(textoInicial,correo_usuario=user_mail)
                        if Texto_Traducido!="None":
                            cell.value=Texto_Traducido
                    except:
                        pass
        #-------HOJA SITE_DUE
        ws1=workbook['SITE DUE']
        columna=column_index_from_string('I')
        fila_ini=11
        fila_fin=37
        for row in ws1.iter_rows(min_row=fila_ini, max_col=columna, max_row=fila_fin,min_col=columna):
            for cell in row:
                textoInicial=cell.value
                if textoInicial!="":
                    #print(textoInicial)
                    #time.sleep(1)
                    try:
                        Texto_Traducido=translate_phrase(textoInicial,correo_usuario=user_mail)
                        if Texto_Traducido!="None":
                            cell.value=Texto_Traducido
                    except:
                        pass
        #-------HOJA ISSUES

        ws1=workbook['ISSUES']
        columna=column_index_from_string('F')
        fila_ini=11
        fila_fin=50
        for row in ws1.iter_rows(min_row=fila_ini, max_col=columna, max_row=fila_fin,min_col=columna):
            for cell in row:
                textoInicial=cell.value
                if textoInicial!="":
                    #print(textoInicial)
                    #time.sleep(1)
                    try:
                        Texto_Traducido=translate_phrase(textoInicial,correo_usuario=user_mail)
                        if Texto_Traducido!="None":
                            cell.value=Texto_Traducido
                    except:
                        pass
        ws1=workbook['ISSUES']
        columna=column_index_from_string('K')
        fila_ini=11
        fila_fin=50
        for row in ws1.iter_rows(min_row=fila_ini, max_col=columna, max_row=fila_fin,min_col=columna):
            for cell in row:
                textoInicial=cell.value
                if textoInicial!="":
                    #print(textoInicial)
                    #time.sleep(1)
                    try:
                        Texto_Traducido=translate_phrase(textoInicial,correo_usuario=user_mail)
                        if Texto_Traducido!="None":
                            cell.value=Texto_Traducido
                    except:
                        pass
        #GUARDARLO EN LA CARPETA DE DESTINO
        workbook.save(directorio+'/Reporte Fotográfico Traducido.xlsx')
        #CERRAR EL DOCUMENTO DE EXCEL
        workbook.close()
        self.lbl_Traduccion.config(text="-Listo-")
MyGUI()